# This script scrapes quotes from Seneca and then sends them via SMS to your phone
import os
from twilio.rest import Client
import openpyxl
import random
import requests
from bs4 import BeautifulSoup
import numpy as np
import pandas as pd


account_sid = "INSERT TWILIO ACCOUNT SID"
auth_token = "INSERT TWILIO AUTH TOKEN"

# Insert file path and desired file name (e.g. /Users/micfranc/Downloads/SenecaQuotes.xlsx
input_drc = 'INSERT FILEPATH WITH DESIRED FILE NAME'

def senecaScrape():

    print("Scraping website for Seneca quotes...")
    url = 'https://www.successories.com/iquote/author/104/lucius-annaeus-seneca-quotes/1'
    headers = {"User-Agent": 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.14; rv:75.0) Gecko/20100101 Firefox/75.0'}
    pages = np.arange(1, 10, 1)
    topics_dict = {"Quote": []}

    for page in pages:
        page = requests.get(url + str(page), headers=headers)
        soup = BeautifulSoup(page.text, 'html.parser')

        for title in soup.findAll("div", {"class": "quote"}):
            quote = title.get_text()
            #print("Page " + str(page))
            topics_dict["Quote"].append(quote)
            topics_data = pd.DataFrame(topics_dict)
            topics_data.to_excel(input_drc, index=False)
            #print(topics_data)


def postFromExcel():

    excel_document = openpyxl.load_workbook(input_drc)
    excel_document.sheetnames
    sheet = excel_document["Sheet1"]
    rand = random.randint(1, 126)
    post = sheet.cell(row=rand, column=1).value
    #print(post)
    client = Client(account_sid, auth_token)
    client.api.account.messages.create(
    to="INSERT TO NUMBER",
    from_="INSERT FROM NUMBER",
    body= post + str("-Seneca the Younger"))

    print("Sent random Seneca quote!")

if __name__ == "__main__":
    senecaScrape()
    postFromExcel()






